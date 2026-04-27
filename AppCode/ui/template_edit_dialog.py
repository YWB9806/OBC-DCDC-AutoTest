"""Excel 模板列映射配置对话框

编辑 Excel 模板的匹配配置和列映射。
"""

import json
import os

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QComboBox, QTableWidget, QTableWidgetItem,
    QPushButton, QGroupBox, QHeaderView, QAbstractItemView,
    QCheckBox, QWidget, QMessageBox
)
from PyQt5.QtCore import Qt


# 数据字段定义
DATA_FIELDS = [
    ("test_result", "测试结果"),
    ("status", "状态"),
    ("duration", "耗时(秒)"),
    ("start_time", "开始时间"),
    ("end_time", "结束时间"),
    ("output", "标准输出"),
    ("error", "错误输出"),
]

MATCH_FIELDS = [
    ("script_name", "脚本名称"),
    ("script_path", "脚本路径"),
]


class TemplateEditDialog(QDialog):
    """Excel 模板列映射配置对话框"""

    def __init__(self, template_path: str, excel_columns: list,
                 config: dict, parent=None):
        """初始化"""
        super().__init__(parent)
        self.template_path = template_path
        self.excel_columns = excel_columns
        self.config = config or {}

        # 读取所有 sheet 名称
        self.sheet_names = self._read_sheet_names()

        self.setWindowTitle(f"匹配映射 - {os.path.basename(template_path)}")
        self.setMinimumSize(800, 650)
        self.setWindowFlag(Qt.WindowMinMaxButtonsHint, True)
        self.resize(900, 720)
        self._init_ui()
        self._load_preview()
        self._load_config()

    def _read_sheet_names(self):
        """读取 Excel 工作簿中的所有 sheet 名称"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.template_path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []

    def _init_ui(self):
        """初始化UI"""
        from PyQt5.QtWidgets import QSplitter, QSpinBox
        layout = QVBoxLayout(self)

        # 模板信息
        info_label = QLabel(f"模板: {os.path.basename(self.template_path)}")
        info_label.setStyleSheet("font-weight: bold; font-size: 11pt;")
        layout.addWidget(info_label)

        # 上下分栏 - 使用 QSplitter 实现可拖拽比例
        splitter = QSplitter(Qt.Vertical)

        # 预览 Excel 前几行
        preview_group = QGroupBox("模板预览")
        preview_layout = QVBoxLayout(preview_group)
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        preview_layout.addWidget(self.preview_table)
        splitter.addWidget(preview_group)

        # 下半部分容器
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        # 匹配配置
        match_group = QGroupBox("匹配配置")
        match_layout = QVBoxLayout(match_group)

        # 第一行：sheet 选择
        sheet_row = QHBoxLayout()
        sheet_row.addWidget(QLabel("映射工作表:"))
        self.sheet_combo = QComboBox()
        for name in self.sheet_names:
            self.sheet_combo.addItem(name)
        sheet_row.addWidget(self.sheet_combo, 1)
        sheet_row.addStretch()
        match_layout.addLayout(sheet_row)

        # 第二行：匹配列和字段
        match_row = QHBoxLayout()
        match_row.addWidget(QLabel("匹配列:"))
        self.match_column_combo = QComboBox()
        for col in self.excel_columns:
            self.match_column_combo.addItem(col)
        match_row.addWidget(self.match_column_combo, 1)

        match_row.addWidget(QLabel("匹配字段:"))
        self.match_field_combo = QComboBox()
        for fk, fl in MATCH_FIELDS:
            self.match_field_combo.addItem(fl, fk)
        match_row.addWidget(self.match_field_combo, 1)

        match_row.addWidget(QLabel("数据起始行:"))
        self.data_start_spin = QSpinBox()
        self.data_start_spin.setMinimum(2)
        self.data_start_spin.setMaximum(100)
        self.data_start_spin.setValue(2)
        match_row.addWidget(self.data_start_spin)
        match_layout.addLayout(match_row)

        bottom_layout.addWidget(match_group)

        # 列映射表
        map_group = QGroupBox("列映射")
        map_layout = QVBoxLayout(map_group)

        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(3)
        self.mapping_table.setHorizontalHeaderLabels(["数据字段", "写入到 Excel 列", "启用"])
        self.mapping_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.mapping_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.mapping_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.mapping_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        map_layout.addWidget(self.mapping_table)

        map_btn_layout = QHBoxLayout()
        add_btn = QPushButton("添加映射")
        add_btn.clicked.connect(self._add_row)
        map_btn_layout.addWidget(add_btn)
        del_btn = QPushButton("删除选中")
        del_btn.clicked.connect(self._delete_row)
        map_btn_layout.addWidget(del_btn)
        map_btn_layout.addStretch()
        map_layout.addLayout(map_btn_layout)

        bottom_layout.addWidget(map_group)

        splitter.addWidget(bottom_widget)

        # 设置初始比例: 预览占1, 下方占2
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        splitter.setSizes([200, 400])

        layout.addWidget(splitter)

        # 底部按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self._on_save)
        btn_layout.addWidget(save_btn)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def _load_preview(self):
        """加载 Excel 预览（前5行）"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.template_path, read_only=True)
            ws = wb.active

            rows_data = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 5:
                    break
                rows_data.append([str(v) if v is not None else '' for v in row])

            wb.close()

            if not rows_data:
                return

            max_cols = max(len(r) for r in rows_data)
            self.preview_table.setColumnCount(max_cols)
            self.preview_table.setRowCount(len(rows_data))

            for r, row_data in enumerate(rows_data):
                for c, val in enumerate(row_data):
                    item = QTableWidgetItem(val)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    if r == 0:
                        item.setBackground(Qt.lightGray)
                    self.preview_table.setItem(r, c, item)

            self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        except Exception as e:
            self.preview_table.setColumnCount(1)
            self.preview_table.setRowCount(1)
            self.preview_table.setItem(0, 0, QTableWidgetItem(f"预览失败: {e}"))

    def _load_config(self):
        """加载配置到 UI"""
        # 选择工作表
        sheet_name = self.config.get('sheet_name', '')
        idx = self.sheet_combo.findText(sheet_name)
        if idx >= 0:
            self.sheet_combo.setCurrentIndex(idx)

        match_col = self.config.get('match_column', '')
        idx = self.match_column_combo.findText(match_col)
        if idx >= 0:
            self.match_column_combo.setCurrentIndex(idx)

        match_field = self.config.get('match_field', 'script_name')
        idx = self.match_field_combo.findData(match_field)
        if idx >= 0:
            self.match_field_combo.setCurrentIndex(idx)

        self.data_start_spin.setValue(self.config.get('data_start_row', 2))

        # 加载列映射
        mappings = self.config.get('column_mappings', [])
        self.mapping_table.setRowCount(len(mappings))

        for row, m in enumerate(mappings):
            self._create_mapping_row(row, m)

    def _create_mapping_row(self, row: int, mapping: dict = None):
        """创建一行映射"""
        mapping = mapping or {}

        # 数据字段下拉
        field_combo = QComboBox()
        for fk, fl in DATA_FIELDS:
            field_combo.addItem(fl, fk)
        if mapping.get('data_field'):
            idx = field_combo.findData(mapping['data_field'])
            if idx >= 0:
                field_combo.setCurrentIndex(idx)
        self.mapping_table.setCellWidget(row, 0, field_combo)

        # Excel 列下拉
        col_combo = QComboBox()
        for c in self.excel_columns:
            col_combo.addItem(c)
        if mapping.get('excel_column'):
            idx = col_combo.findText(mapping['excel_column'])
            if idx >= 0:
                col_combo.setCurrentIndex(idx)
        self.mapping_table.setCellWidget(row, 1, col_combo)

        # 启用复选框
        checkbox = QCheckBox()
        checkbox.setChecked(mapping.get('enabled', True))
        cell_widget = QWidget()
        cb_layout = QHBoxLayout(cell_widget)
        cb_layout.addWidget(checkbox)
        cb_layout.setAlignment(Qt.AlignCenter)
        cb_layout.setContentsMargins(0, 0, 0, 0)
        self.mapping_table.setCellWidget(row, 2, cell_widget)

    def _add_row(self):
        """添加映射行"""
        row = self.mapping_table.rowCount()
        self.mapping_table.insertRow(row)
        self._create_mapping_row(row)

    def _delete_row(self):
        """删除选中的映射行"""
        rows = set()
        for item in self.mapping_table.selectedItems():
            rows.add(item.row())
        for row in sorted(rows, reverse=True):
            self.mapping_table.removeRow(row)

    def _on_save(self):
        """保存配置"""
        match_col = self.match_column_combo.currentText()
        if not match_col:
            QMessageBox.warning(self, "警告", "请选择匹配列")
            return

        self.accept()

    def get_config_data(self) -> dict:
        """获取配置数据"""
        mappings = []
        for row in range(self.mapping_table.rowCount()):
            field_combo = self.mapping_table.cellWidget(row, 0)
            col_combo = self.mapping_table.cellWidget(row, 1)
            if not field_combo or not col_combo:
                continue

            cell_widget = self.mapping_table.cellWidget(row, 2)
            enabled = True
            if cell_widget:
                cb = cell_widget.findChild(QCheckBox)
                if cb:
                    enabled = cb.isChecked()

            mappings.append({
                'data_field': field_combo.currentData(),
                'excel_column': col_combo.currentText(),
                'enabled': enabled,
            })

        return {
            'sheet_name': self.sheet_combo.currentText() if self.sheet_combo.count() > 0 else '',
            'match_column': self.match_column_combo.currentText(),
            'match_field': self.match_field_combo.currentData(),
            'header_row': 1,
            'data_start_row': self.data_start_spin.value(),
            'column_mappings': mappings,
            'excel_columns': self.excel_columns,
        }
