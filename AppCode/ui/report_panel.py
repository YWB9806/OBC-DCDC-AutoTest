"""报告面板

提供 Excel 模板导入、列映射配置和报告生成功能。
"""

import json
import os
import shutil
from datetime import datetime
from typing import Dict, Any, List, Optional

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QComboBox, QPushButton, QLabel, QTextBrowser,
    QFileDialog, QMessageBox, QDateEdit,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView
)
from PyQt5.QtCore import Qt, QDate

from AppCode.utils.constants import TestResult


EXCEL_TEMPLATE_DIR = os.path.join("data", "report_templates", "excel")

# 数据字段定义
DATA_FIELDS = [
    ("test_result", "测试结果"),
    ("status", "状态"),
    ("duration", "耗时(秒)"),
    ("start_time", "开始时间"),
    ("end_time", "结束时间"),
    ("output", "标准输出"),
    ("error", "错误输出"),
]

# 匹配字段定义
MATCH_FIELDS = [
    ("script_name", "脚本名称"),
    ("script_path", "脚本路径"),
]


class ReportPanel(QWidget):
    """报告面板"""

    def __init__(self, container, parent=None):
        super().__init__(parent)
        self.container = container
        self.logger = container.resolve('log_manager').get_logger('report_panel')
        self.history_repo = container.resolve('execution_history_repo')
        self.test_suite_repo = container.resolve('test_suite_repository')

        self._excel_columns = []  # 当前模板的 Excel 列名列表
        self._cached_records = []  # 加载的执行记录缓存
        self._record_map = {}  # 匹配键 -> 记录

        self._init_ui()
        self._load_templates()

    def _init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(4)

        # === 1. 模板管理区 ===
        template_group = QGroupBox("Excel 模板")
        template_layout = QHBoxLayout(template_group)
        template_layout.setContentsMargins(6, 6, 6, 6)

        template_layout.addWidget(QLabel("选择模板:"))
        self.template_combo = QComboBox()
        self.template_combo.currentIndexChanged.connect(self._on_template_changed)
        template_layout.addWidget(self.template_combo, 1)

        import_btn = QPushButton("导入模板")
        import_btn.clicked.connect(self._on_import_template)
        template_layout.addWidget(import_btn)

        config_btn = QPushButton("配置映射")
        config_btn.clicked.connect(self._on_config_mapping)
        template_layout.addWidget(config_btn)

        del_btn = QPushButton("删除模板")
        del_btn.clicked.connect(self._on_delete_template)
        template_layout.addWidget(del_btn)

        layout.addWidget(template_group)

        # === 2. 生成报告（紧凑布局） ===
        gen_group = QGroupBox("生成报告")
        gen_layout = QVBoxLayout(gen_group)
        gen_layout.setContentsMargins(6, 6, 6, 6)
        gen_layout.setSpacing(4)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("方案:"))
        self.suite_combo = QComboBox()
        self.suite_combo.addItem("全部方案", "")
        self.suite_combo.currentIndexChanged.connect(self._load_batches)
        row1.addWidget(self.suite_combo, 1)

        row1.addWidget(QLabel("批次:"))
        self.batch_combo = QComboBox()
        self.batch_combo.addItem("全部批次", "")
        row1.addWidget(self.batch_combo, 1)

        row1.addWidget(QLabel("日期:"))
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate().addMonths(-1))
        self.start_date.dateChanged.connect(self._load_batches)
        row1.addWidget(self.start_date)

        row1.addWidget(QLabel("至"))
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())
        self.end_date.dateChanged.connect(self._load_batches)
        row1.addWidget(self.end_date)

        gen_layout.addLayout(row1)

        btn_layout = QHBoxLayout()
        refresh_btn = QPushButton("刷新方案")
        refresh_btn.clicked.connect(self._load_suites)
        btn_layout.addWidget(refresh_btn)

        load_data_btn = QPushButton("加载数据")
        load_data_btn.clicked.connect(self._on_load_data)
        btn_layout.addWidget(load_data_btn)

        preview_btn = QPushButton("生成预览")
        preview_btn.clicked.connect(self._on_preview)
        btn_layout.addWidget(preview_btn)

        export_btn = QPushButton("导出 Excel")
        export_btn.clicked.connect(self._on_export_excel)
        btn_layout.addWidget(export_btn)

        # 状态标签放在按钮行末尾
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #666; padding: 2px;")
        btn_layout.addWidget(self.status_label, 1)
        gen_layout.addLayout(btn_layout)

        layout.addWidget(gen_group)

        # === 4. 预览区 ===
        preview_group = QGroupBox("报告预览")
        preview_layout = QVBoxLayout(preview_group)
        preview_layout.setContentsMargins(6, 6, 6, 6)
        self.preview_browser = QTextBrowser()
        self.preview_browser.setOpenExternalLinks(False)
        preview_layout.addWidget(self.preview_browser)
        layout.addWidget(preview_group, 1)

        self._load_suites()

    # ==================== 模板管理 ====================

    def _load_templates(self):
        """加载 Excel 模板列表"""
        self.template_combo.clear()
        os.makedirs(EXCEL_TEMPLATE_DIR, exist_ok=True)

        for fname in os.listdir(EXCEL_TEMPLATE_DIR):
            if fname.endswith('.xlsx') and not fname.startswith('~'):
                self.template_combo.addItem(fname)

    def _get_template_path(self) -> Optional[str]:
        """获取当前选中模板的文件路径"""
        name = self.template_combo.currentText()
        if not name:
            return None
        path = os.path.join(EXCEL_TEMPLATE_DIR, name)
        return path if os.path.exists(path) else None

    def _get_config_path(self) -> Optional[str]:
        """获取当前选中模板的配置文件路径"""
        name = self.template_combo.currentText()
        if not name:
            return None
        base = os.path.splitext(name)[0]
        return os.path.join(EXCEL_TEMPLATE_DIR, f"{base}.json")

    def _on_import_template(self):
        """导入 Excel 模板"""
        src_path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 模板", "",
            "Excel Files (*.xlsx)"
        )
        if not src_path:
            return

        try:
            fname = os.path.basename(src_path)
            dst_path = os.path.join(EXCEL_TEMPLATE_DIR, fname)

            if os.path.exists(dst_path):
                reply = QMessageBox.question(
                    self, "确认覆盖",
                    f"模板「{fname}」已存在，是否覆盖？",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return

            shutil.copy2(src_path, dst_path)

            # 读取表头行，初始化默认配置
            excel_columns = self._read_excel_columns(dst_path)
            if excel_columns:
                default_config = {
                    "sheet_name": "",
                    "match_column": excel_columns[0],
                    "match_field": "script_name",
                    "header_row": 1,
                    "data_start_row": 2,
                    "column_mappings": [],
                    "excel_columns": excel_columns,
                }
                config_path = os.path.join(
                    EXCEL_TEMPLATE_DIR,
                    f"{os.path.splitext(fname)[0]}.json"
                )
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(default_config, f, ensure_ascii=False, indent=2)

            self._load_templates()
            idx = self.template_combo.findText(fname)
            if idx >= 0:
                self.template_combo.setCurrentIndex(idx)

            self.logger.info(f"Imported template: {fname}")
            QMessageBox.information(self, "成功", f"模板已导入:\n{fname}")
        except Exception as e:
            self.logger.error(f"Import template failed: {e}")
            QMessageBox.critical(self, "错误", f"导入模板失败:\n{e}")

    def _on_delete_template(self):
        """删除当前模板"""
        name = self.template_combo.currentText()
        if not name:
            return

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除模板「{name}」及其配置吗？",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            xlsx_path = os.path.join(EXCEL_TEMPLATE_DIR, name)
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            config_path = self._get_config_path()
            if config_path and os.path.exists(config_path):
                os.remove(config_path)
            self._load_templates()

    def _on_template_changed(self):
        """模板切换 — 加载配置"""
        path = self._get_template_path()
        if not path:
            return

        self._excel_columns = self._read_excel_columns(path)
        self._load_mapping_config()

    def _read_excel_columns(self, xlsx_path: str) -> List[str]:
        """读取 Excel 文件的表头列名"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            ws = wb.active

            config_path = os.path.join(
                EXCEL_TEMPLATE_DIR,
                f"{os.path.splitext(os.path.basename(xlsx_path))[0]}.json"
            )
            header_row = 1
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                header_row = cfg.get('header_row', 1)

            columns = []
            for cell in ws[header_row]:
                if cell.value is not None:
                    columns.append(str(cell.value).strip())
                else:
                    break

            wb.close()
            return columns
        except Exception as e:
            self.logger.error(f"Failed to read Excel columns: {e}")
            return []

    def _on_config_mapping(self):
        """打开完整映射配置对话框"""
        template_path = self._get_template_path()
        if not template_path:
            QMessageBox.warning(self, "警告", "请先选择一个模板")
            return

        from AppCode.ui.template_edit_dialog import TemplateEditDialog

        config = self._load_config_dict()
        excel_columns = self._excel_columns

        dialog = TemplateEditDialog(template_path, excel_columns, config, self)
        if dialog.exec_():
            new_config = dialog.get_config_data()
            config_path = self._get_config_path()
            if config_path:
                os.makedirs(os.path.dirname(config_path), exist_ok=True)
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(new_config, f, ensure_ascii=False, indent=2)
                self.logger.info(f"Saved mapping config: {config_path}")
                self._load_mapping_config()

    # ==================== 配置管理 ====================

    def _load_config_dict(self) -> dict:
        """加载配置字典"""
        config_path = self._get_config_path()
        if config_path and os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                self.logger.error(f"Failed to load config: {e}")
        return {
            "match_column": "",
            "match_field": "script_name",
            "header_row": 1,
            "data_start_row": 2,
            "column_mappings": [],
            "excel_columns": self._excel_columns,
        }

    def _load_mapping_config(self):
        """加载映射配置（配置映射通过对话框管理，此处仅缓存配置）"""
        # 配置由 _on_config_mapping 对话框管理，无需更新已移除的UI控件
        pass

    def _populate_mapping_table(self, mappings: list):
        """填充映射表"""
        self.mapping_table.setRowCount(len(mappings))
        for row, m in enumerate(mappings):
            # 数据字段下拉
            field_combo = QComboBox()
            for fk, fl in DATA_FIELDS:
                field_combo.addItem(fl, fk)
            idx = field_combo.findData(m.get('data_field', ''))
            if idx >= 0:
                field_combo.setCurrentIndex(idx)
            self.mapping_table.setCellWidget(row, 0, field_combo)

            # Excel 列下拉
            col_combo = QComboBox()
            for c in self._excel_columns:
                col_combo.addItem(c)
            idx = col_combo.findText(m.get('excel_column', ''))
            if idx >= 0:
                col_combo.setCurrentIndex(idx)
            self.mapping_table.setCellWidget(row, 1, col_combo)

            # 启用复选框
            from PyQt5.QtWidgets import QCheckBox
            checkbox = QCheckBox()
            checkbox.setChecked(m.get('enabled', True))
            cell_widget = QWidget()
            cb_layout = QHBoxLayout(cell_widget)
            cb_layout.addWidget(checkbox)
            cb_layout.setAlignment(Qt.AlignCenter)
            cb_layout.setContentsMargins(0, 0, 0, 0)
            self.mapping_table.setCellWidget(row, 2, cell_widget)

    def _add_mapping_row(self):
        """添加映射行"""
        row = self.mapping_table.rowCount()
        self.mapping_table.insertRow(row)

        # 数据字段下拉
        field_combo = QComboBox()
        for fk, fl in DATA_FIELDS:
            field_combo.addItem(fl, fk)
        self.mapping_table.setCellWidget(row, 0, field_combo)

        # Excel 列下拉
        col_combo = QComboBox()
        for c in self._excel_columns:
            col_combo.addItem(c)
        self.mapping_table.setCellWidget(row, 1, col_combo)

        # 启用复选框
        from PyQt5.QtWidgets import QCheckBox
        checkbox = QCheckBox()
        checkbox.setChecked(True)
        cell_widget = QWidget()
        cb_layout = QHBoxLayout(cell_widget)
        cb_layout.addWidget(checkbox)
        cb_layout.setAlignment(Qt.AlignCenter)
        cb_layout.setContentsMargins(0, 0, 0, 0)
        self.mapping_table.setCellWidget(row, 2, cell_widget)

    def _delete_mapping_row(self):
        """删除选中的映射行"""
        rows = set()
        for item in self.mapping_table.selectedItems():
            rows.add(item.row())
        for row in sorted(rows, reverse=True):
            self.mapping_table.removeRow(row)

    def _get_mappings_from_table(self) -> list:
        """从映射表读取配置"""
        mappings = []
        for row in range(self.mapping_table.rowCount()):
            field_combo = self.mapping_table.cellWidget(row, 0)
            col_combo = self.mapping_table.cellWidget(row, 1)
            if not field_combo or not col_combo:
                continue

            # 获取复选框状态
            cell_widget = self.mapping_table.cellWidget(row, 2)
            enabled = True
            if cell_widget:
                cb = cell_widget.findChild(QCheckBox)
                if cb:
                    enabled = cb.isChecked()

            mappings.append({
                'data_field': field_combo.currentData(),
                'excel_column': col_combo.currentText(),
                'enabled': enabled,
            })
        return mappings

    def _save_mapping_config(self):
        """保存映射配置"""
        config = {
            "match_column": self.match_column_combo.currentText(),
            "match_field": self.match_field_combo.currentData(),
            "header_row": 1,
            "data_start_row": 2,
            "column_mappings": self._get_mappings_from_table(),
            "excel_columns": self._excel_columns,
        }
        config_path = self._get_config_path()
        if config_path:
            os.makedirs(os.path.dirname(config_path), exist_ok=True)
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            self.logger.info("Mapping config saved")

    # ==================== 数据加载与匹配 ====================

    def _load_suites(self):
        """加载测试方案列表"""
        self.suite_combo.blockSignals(True)
        self.suite_combo.clear()
        self.suite_combo.addItem("全部方案", "")
        try:
            suites = self.test_suite_repo.get_all()
            for suite in suites:
                self.suite_combo.addItem(suite.get('name', ''), suite.get('id'))
        except Exception as e:
            self.logger.error(f"Failed to load suites: {e}")
        self.suite_combo.blockSignals(False)
        self._load_batches()

    def _get_report_data(self, batch_id: str = '') -> List[Dict[str, Any]]:
        """获取报告数据（按日期范围、方案和批次过滤）"""
        suite_name = self.suite_combo.currentText()
        start = self.start_date.date().toString("yyyy-MM-dd")
        end = self.end_date.date().toString("yyyy-MM-dd")

        records = self.history_repo.get_by_date_range(start, end)

        if suite_name and suite_name != "全部方案":
            records = [r for r in records if r.get('suite_name') == suite_name]
        if batch_id:
            records = [r for r in records if r.get('batch_id') == batch_id]

        return records

    def _load_batches(self):
        """根据日期和方案刷新批次列表"""
        self.batch_combo.blockSignals(True)
        self.batch_combo.clear()
        self.batch_combo.addItem("全部批次", "")

        records = self._get_report_data()
        batch_map = {}
        for r in records:
            bid = r.get('batch_id', '')
            if bid:
                batch_map[bid] = batch_map.get(bid, 0) + 1

        # 按时间倒序排列
        sorted_batches = sorted(batch_map.items(), key=lambda x: x[0], reverse=True)
        for bid, count in sorted_batches:
            label = bid
            try:
                parts = bid.split('_')
                ts = int(parts[1]) / 1000000
                dt = datetime.fromtimestamp(ts)
                label = f"{dt.strftime('%m-%d %H:%M:%S')} ({count}条)"
            except Exception:
                pass
            self.batch_combo.addItem(label, bid)

        self.batch_combo.blockSignals(False)

    def _on_load_data(self):
        """加载数据 — 查询执行记录并建立匹配映射"""
        try:
            batch_id = self.batch_combo.currentData() or ''
            records = self._get_report_data(batch_id)
            self._cached_records = records

            # 建立匹配映射（从配置获取匹配字段，而非已移除的UI控件）
            config = self._load_config_dict()
            match_field = config.get('match_field', 'script_name')
            self._record_map = {}
            for r in records:
                key = r.get(match_field, '')
                if match_field == 'script_name':
                    key = os.path.splitext(os.path.basename(r.get('script_path', '')))[0]
                if key:
                    key = str(key).strip()
                    self._record_map[key] = r
                    # 同时保留带扩展名的键，兼容 Excel 中写 .py 的情况
                    if match_field == 'script_name':
                        full_name = os.path.basename(r.get('script_path', '')).strip()
                        if full_name and full_name != key:
                            self._record_map[full_name] = r

            batch_label = self.batch_combo.currentText() if batch_id else "全部"
            self.status_label.setText(
                f"批次: {batch_label} | "
                f"已加载 {len(records)} 条记录，"
                f"匹配键 {len(self._record_map)} 个"
            )
            self.logger.info(f"Loaded {len(records)} records, {len(self._record_map)} match keys")
        except Exception as e:
            self.logger.error(f"Load data failed: {e}")
            QMessageBox.critical(self, "错误", f"加载数据失败:\n{e}")

    def _on_preview(self):
        """生成预览 — 显示填充数据后的报告表格"""
        template_path = self._get_template_path()
        if not template_path:
            QMessageBox.warning(self, "警告", "请先选择一个 Excel 模板")
            return

        if not self._cached_records:
            self._on_load_data()

        if not self._cached_records:
            self.preview_browser.setHtml("<p style='color:#999;'>所选时间范围内没有执行记录</p>")
            return

        config = self._load_config_dict()
        match_col = config.get('match_column', '')
        mappings = config.get('column_mappings', [])

        if not match_col:
            QMessageBox.warning(self, "警告", "请先配置匹配列")
            return

        enabled_mappings = [m for m in mappings if m.get('enabled')]
        if not enabled_mappings:
            QMessageBox.warning(self, "警告", "请先配置至少一个启用的列映射")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(template_path, read_only=True)
            ws = self._get_worksheet(wb, config)

            header_row = config.get('header_row', 1)
            data_start_row = config.get('data_start_row', 2)

            # 找到匹配列索引
            match_col_idx = self._find_column_index(ws, match_col, header_row)
            if match_col_idx is None:
                wb.close()
                QMessageBox.warning(self, "警告", f"在模板中找不到匹配列「{match_col}」")
                return

            # 构建表头
            headers = [match_col] + [m['excel_column'] for m in enabled_mappings]

            # 遍历数据行，收集匹配数据
            rows_data = []
            matched = 0
            total = 0
            stats = {'pass': 0, 'fail': 0, 'pending': 0, 'error': 0, 'timeout': 0}

            for row in ws.iter_rows(min_row=data_start_row):
                match_val = row[match_col_idx - 1].value
                if match_val is None:
                    continue
                total += 1
                match_key = str(match_val).strip()
                record = self._record_map.get(match_key)

                row_cells = [match_key]
                if record:
                    matched += 1
                    for m in enabled_mappings:
                        val = self._format_field_value(record, m['data_field'])
                        row_cells.append(str(val) if val is not None else '')
                    tr = record.get('test_result', '')
                    if tr in stats:
                        stats[tr] += 1
                else:
                    for m in enabled_mappings:
                        row_cells.append('-')
                rows_data.append(row_cells)

            wb.close()

            # 生成 HTML
            html = "<h3>测试报告预览</h3>"
            html += "<p>模板: {} | ".format(os.path.basename(template_path))
            batch_label = self.batch_combo.currentText() if self.batch_combo.currentData() else "全部"
            html += "批次: {} | ".format(batch_label)
            html += "匹配: <b style='color:green'>{}</b>/{}</p>".format(matched, total)

            # HTML 表格
            html += "<table border='1' cellpadding='4' cellspacing='0' style='border-collapse:collapse;'>"
            html += "<tr style='background:#e0e0e0;'>"
            for h in headers:
                html += "<th>{}</th>".format(h)
            html += "</tr>"
            for row_cells in rows_data[:100]:
                html += "<tr>"
                for i, val in enumerate(row_cells):
                    if i > 0 and val == '-':
                        html += "<td style='color:#999;'>-</td>"
                    else:
                        display = val[:100] + ('...' if len(val) > 100 else '') if len(val) > 100 else val
                        html += "<td>{}</td>".format(display)
                html += "</tr>"
            html += "</table>"

            if len(rows_data) > 100:
                html += "<p style='color:#999;'>... 共 {} 行，仅显示前 100 行</p>".format(len(rows_data))

            # 统计信息
            total_result = sum(stats.values())
            if total_result > 0:
                pass_rate = round(stats['pass'] / total_result * 100, 1)
                html += "<p>统计: 通过 <b style='color:green'>{}</b>".format(stats['pass'])
                html += " | 失败 <b style='color:red'>{}</b>".format(stats['fail'])
                html += " | 待判定 {}".format(stats['pending'])
                html += " | 错误 {}".format(stats['error'])
                html += " | 超时 {}".format(stats['timeout'])
                html += " | 通过率 <b>{}%</b></p>".format(pass_rate)

            self.preview_browser.setHtml(html)
            self.logger.info(f"Preview: {matched}/{total} matched")

        except Exception as e:
            self.logger.error(f"Preview failed: {e}")
            QMessageBox.critical(self, "错误", f"生成预览失败:\n{e}")

    # ==================== 导出 ====================

    def _on_export_excel(self):
        """导出 Excel — 基于模板匹配写入"""
        template_path = self._get_template_path()
        if not template_path:
            QMessageBox.warning(self, "警告", "请先选择一个 Excel 模板")
            return

        config = self._load_config_dict()
        match_col = config.get('match_column', '')
        mappings = [m for m in config.get('column_mappings', []) if m.get('enabled')]

        if not match_col:
            QMessageBox.warning(self, "警告", "请先配置匹配列")
            return

        if not mappings:
            QMessageBox.warning(self, "警告", "请先配置至少一个列映射")
            return

        if not self._cached_records:
            self._on_load_data()

        if not self._cached_records:
            QMessageBox.information(self, "提示", "所选时间范围内没有执行记录")
            return

        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.critical(
                self, "缺少依赖",
                "需要 openpyxl 库，请运行:\npip install openpyxl"
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出报告",
            f"测试报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(template_path)
            ws = self._get_worksheet(wb, config)

            header_row = config.get('header_row', 1)
            data_start_row = config.get('data_start_row', 2)

            # 找到匹配列索引
            match_col_idx = self._find_column_index(ws, match_col, header_row)
            if match_col_idx is None:
                QMessageBox.warning(self, "警告", f"在模板中找不到匹配列「{match_col}」")
                return

            # 找到各映射目标列索引
            col_indices = {}
            for m in mappings:
                col_idx = self._find_column_index(ws, m['excel_column'], header_row)
                if col_idx is not None:
                    col_indices[m['data_field']] = col_idx

            # 匹配字段
            match_field = config.get('match_field', 'script_name')

            # 遍历 Excel 数据行并填入
            filled = 0
            for row in ws.iter_rows(min_row=data_start_row):
                match_val = row[match_col_idx - 1].value
                if match_val is None:
                    continue
                match_key = str(match_val).strip()
                record = self._record_map.get(match_key)
                if not record:
                    continue

                for data_field, col_idx in col_indices.items():
                    val = self._format_field_value(record, data_field)
                    ws.cell(row=row[0].row, column=col_idx).value = val

                filled += 1

            wb.save(file_path)
            QMessageBox.information(
                self, "成功",
                f"报告已导出到:\n{file_path}\n\n匹配写入: {filled} 行"
            )
            self.logger.info(f"Excel exported: {file_path}, filled {filled} rows")

        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            QMessageBox.critical(self, "错误", f"导出 Excel 失败:\n{e}")

    def _get_worksheet(self, wb, config: dict):
        """根据配置获取要映射的工作表"""
        sheet_name = config.get('sheet_name', '')
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.active

    def _find_column_index(self, ws, column_name: str, header_row: int) -> Optional[int]:
        """在工作表中查找列名对应的列索引（1-based）"""
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip() == column_name:
                return cell.column
        return None

    def _format_field_value(self, record: dict, field: str) -> str:
        """格式化字段值，用于写入 Excel"""
        val = record.get(field, '')
        if field == 'test_result':
            # 将英文结果转为中文显示
            result_map = {
                'pass': TestResult.PASS,
                'fail': TestResult.FAIL,
                'pending': TestResult.PENDING,
                'error': TestResult.ERROR,
                'timeout': TestResult.TIMEOUT,
            }
            val = result_map.get(str(val), val)
        elif field == 'duration':
            try:
                val = round(float(val), 2)
            except (ValueError, TypeError):
                pass
        return val if val is not None else ''

    def refresh(self):
        """刷新面板"""
        self._load_suites()
        self._load_templates()
